"""Native LightRAG document ingestion helpers."""

from __future__ import annotations

import logging
import time
import uuid
from itertools import zip_longest
from pathlib import Path
from typing import Any

from lightrag.constants import FULL_DOCS_FORMAT_PENDING_PARSE, FULL_DOCS_FORMAT_RAW
from lightrag.parser.routing import resolve_file_parser_directives
from lightrag.utils import compute_mdhash_id

from src.server.document_processing import record_failed_doc

logger = logging.getLogger(__name__)


TEXT_BEARING_TABLE_SUFFIXES = {
    ".csv",
    ".doc",
    ".docx",
    ".md",
    ".txt",
    ".xls",
    ".xlsm",
    ".xlsx",
    ".tsv",
}

OPENPYXL_WORKBOOK_SUFFIXES = {".xlsx", ".xlsm"}
SPREADSHEET_SUFFIXES = {".xls", *OPENPYXL_WORKBOOK_SUFFIXES}
MAX_SPREADSHEET_CELLS = 50_000


def _new_track_id(file_name: str) -> str:
    stem = Path(file_name).stem or "document"
    safe_stem = "".join(ch if ch.isalnum() or ch in "-_" else "-" for ch in stem)[:48]
    return f"native-{safe_stem}-{uuid.uuid4().hex[:8]}"


def _suppress_text_bearing_table_analysis(file_name: str, process_options: str) -> str:
    suffix = Path(file_name).suffix.lower()
    if suffix not in TEXT_BEARING_TABLE_SUFFIXES:
        return process_options
    return "".join(option for option in process_options if option != "t")


def resolve_govcon_parser_directives(file_name: str) -> tuple[str, str]:
    if Path(file_name).suffix.lower() in SPREADSHEET_SUFFIXES:
        return "native", ""
    parser_engine, process_options = resolve_file_parser_directives(
        file_name,
        require_external_endpoint=False,
    )
    return parser_engine, _suppress_text_bearing_table_analysis(file_name, process_options)


def _is_spreadsheet(file_name: str) -> bool:
    return Path(file_name).suffix.lower() in SPREADSHEET_SUFFIXES


def _format_spreadsheet_value(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", " ").replace("\n", " ").strip()
    return " ".join(text.split())


def _extract_spreadsheet_text(file_path: str, file_name: str) -> str:
    suffix = Path(file_name).suffix.lower()
    if suffix not in OPENPYXL_WORKBOOK_SUFFIXES:
        raise ValueError(
            "Native spreadsheet extraction supports .xlsx/.xlsm only; "
            f"convert {file_name} to .xlsx."
        )

    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is declared in project env
        raise RuntimeError(
            "openpyxl is required for native spreadsheet extraction"
        ) from exc

    values_workbook = load_workbook(file_path, read_only=True, data_only=True)
    formulas_workbook = load_workbook(file_path, read_only=True, data_only=False)
    lines = [f"# Workbook: {file_name}"]
    cell_count = 0
    truncated = False
    try:
        for sheet_name in values_workbook.sheetnames:
            values_sheet = values_workbook[sheet_name]
            formulas_sheet = formulas_workbook[sheet_name]
            lines.append("")
            lines.append(f"## Sheet: {sheet_name}")
            lines.append(
                f"Dimensions: {values_sheet.max_row} rows x "
                f"{values_sheet.max_column} columns"
            )
            non_empty_rows = 0

            for values_row, formulas_row in zip_longest(
                values_sheet.iter_rows(), formulas_sheet.iter_rows(), fillvalue=()
            ):
                row_cells = []
                row_number = None
                for value_cell, formula_cell in zip_longest(values_row, formulas_row):
                    cell = value_cell or formula_cell
                    if cell is None:
                        continue
                    row_number = row_number or getattr(cell, "row", None)
                    value_text = _format_spreadsheet_value(
                        getattr(value_cell, "value", None)
                    )
                    formula_value = getattr(formula_cell, "value", None)
                    formula_text = _format_spreadsheet_value(formula_value)
                    if not value_text and not formula_text:
                        continue

                    display = value_text
                    if formula_text.startswith("="):
                        display = (
                            f"{value_text} (formula: {formula_text})"
                            if value_text
                            else formula_text
                        )
                    coordinate = getattr(cell, "coordinate", "cell")
                    row_cells.append(f"{coordinate}={display}")
                    cell_count += 1
                    if cell_count >= MAX_SPREADSHEET_CELLS:
                        truncated = True
                        break
                if row_cells:
                    non_empty_rows += 1
                    lines.append(
                        f"Row {row_number or non_empty_rows}: "
                        + " | ".join(row_cells)
                    )
                if truncated:
                    break

            if non_empty_rows == 0:
                lines.append("No non-empty cells detected.")
            if truncated:
                lines.append(
                    f"[TRUNCATED after {MAX_SPREADSHEET_CELLS} non-empty cells]"
                )
                break
    finally:
        values_workbook.close()
        formulas_workbook.close()

    return "\n".join(lines).strip() + "\n"


def _status_value(status: Any) -> str:
    return str(getattr(status, "value", status) or "").lower()


def _is_retryable_failed_status(record: dict[str, Any]) -> bool:
    return _status_value(record.get("status")) == "failed"


def _status_field(record: Any, field_name: str) -> Any:
    if isinstance(record, dict):
        return record.get(field_name)
    return getattr(record, field_name, None)


def _is_duplicate_status(record: Any) -> bool:
    metadata = _status_field(record, "metadata")
    if not isinstance(metadata, dict):
        metadata = {}
    summary = str(_status_field(record, "content_summary") or "")
    return bool(metadata.get("is_duplicate")) or summary.startswith("[DUPLICATE:")


async def _matching_failed_status_ids(doc_status: Any, file_name: str) -> list[str]:
    storage_data = getattr(doc_status, "_data", None)
    storage_lock = getattr(doc_status, "_storage_lock", None)
    if storage_data is not None:
        records: list[tuple[str, dict[str, Any]]] = []
        if storage_lock is None:
            records = list(storage_data.items())
        else:
            async with storage_lock:
                records = list(storage_data.items())
        return [
            doc_id
            for doc_id, record in records
            if isinstance(record, dict)
            and record.get("file_path") == file_name
            and _is_retryable_failed_status(record)
        ]

    get_by_basename = getattr(doc_status, "get_doc_by_file_basename", None)
    if get_by_basename is None:
        return []
    match = await get_by_basename(file_name)
    if not match:
        return []
    doc_id, record = match
    if isinstance(record, dict) and _is_retryable_failed_status(record):
        return [doc_id]
    return []


async def _clear_retryable_failed_statuses(lightrag: Any, file_name: str) -> list[str]:
    doc_status = getattr(lightrag, "doc_status", None)
    delete_status = getattr(doc_status, "delete", None)
    if doc_status is None or delete_status is None:
        return []

    failed_ids = await _matching_failed_status_ids(doc_status, file_name)
    if not failed_ids:
        return []

    await delete_status(failed_ids)
    index_done = getattr(doc_status, "index_done_callback", None)
    if index_done is not None:
        await index_done()
    logger.info(
        "Cleared %s retryable failed native doc_status record(s) for %s",
        len(failed_ids),
        file_name,
    )
    return failed_ids


async def _native_track_failures(lightrag: Any, track_id: str) -> list[tuple[str, str]]:
    doc_status = getattr(lightrag, "doc_status", None)
    get_by_track = getattr(doc_status, "get_docs_by_track_id", None)
    if get_by_track is None:
        return []

    track_docs = await get_by_track(track_id)
    failures: list[tuple[str, str]] = []
    for doc_id, record in (track_docs or {}).items():
        if _status_value(_status_field(record, "status")) != "failed":
            continue
        if _is_duplicate_status(record):
            continue
        error = str(_status_field(record, "error_msg") or "native pipeline failed")
        failures.append((doc_id, error))
    return failures


async def process_document_with_native_ingestion(
    file_path: str,
    file_name: str,
    rag_instance: Any,
    llm_func: Any | None = None,
    *,
    track_id: str | None = None,
    from_scan: bool = False,
    callback: Any | None = None,
) -> dict[str, Any]:
    """Queue one source document through LightRAG's native parser pipeline."""

    del llm_func
    lightrag = rag_instance.lightrag
    resolved_track_id = track_id or _new_track_id(file_name)
    parser_engine, process_options = resolve_govcon_parser_directives(file_name or file_path)
    doc_id = compute_mdhash_id(Path(file_name or file_path).name, prefix="doc-")
    start_time = time.perf_counter()

    try:
        logger.info(
            "📄 Native LightRAG ingest %s (engine=%s options=%s track=%s)",
            file_name,
            parser_engine,
            process_options or "default",
            resolved_track_id,
        )
        source_name = Path(file_name or file_path).name
        await _clear_retryable_failed_statuses(lightrag, source_name)
        if _is_spreadsheet(source_name):
            spreadsheet_text = _extract_spreadsheet_text(file_path, source_name)
            await lightrag.apipeline_enqueue_documents(
                spreadsheet_text,
                file_paths=file_path,
                track_id=resolved_track_id,
                docs_format=FULL_DOCS_FORMAT_RAW,
                parse_engine=parser_engine,
                process_options=process_options,
                from_scan=from_scan,
            )
        else:
            await lightrag.apipeline_enqueue_documents(
                "",
                file_paths=file_path,
                track_id=resolved_track_id,
                docs_format=FULL_DOCS_FORMAT_PENDING_PARSE,
                parse_engine=parser_engine,
                process_options=process_options,
                from_scan=from_scan,
            )
        await lightrag.apipeline_process_enqueue_documents()
        failures = await _native_track_failures(lightrag, resolved_track_id)
        if failures:
            detail = "; ".join(f"{doc_id}: {error}" for doc_id, error in failures)
            raise RuntimeError(f"Native LightRAG pipeline failed for {file_name}: {detail}")
        if callback is not None:
            callback.on_document_complete(
                file_path=file_path,
                doc_id=doc_id,
                duration_seconds=time.perf_counter() - start_time,
            )
        return {
            "status": "success",
            "relationships_inferred": 0,
            "method": "native_lightrag_pipeline",
            "message": "Document queued and processed by LightRAG native pipeline.",
            "track_id": resolved_track_id,
            "workspace": getattr(lightrag, "workspace", ""),
        }
    except Exception as exc:
        await record_failed_doc(
            rag_instance,
            file_path,
            Path(file_name or file_path).name,
            doc_id,
            str(exc),
        )
        if callback is not None:
            callback.on_document_error(
                file_path=file_path,
                doc_id=doc_id,
                error=str(exc),
            )
        raise
