from __future__ import annotations

import json
import os
import subprocess
import sys
import zipfile
from pathlib import Path


REPO_ROOT = Path(__file__).resolve().parents[2]
RENDERERS = REPO_ROOT / ".github" / "skills" / "renderers" / "scripts"


def test_render_xlsx_discovers_nested_array_sheets(tmp_path: Path) -> None:
    source = tmp_path / "competitive_intel_obligation.json"
    output = tmp_path / "competitive_intel_obligation.xlsx"
    source.write_text(
        json.dumps(
            {
                "obligations": {
                    "by_award": [
                        {"award_id": "FA805122F0001", "obligated": 1000},
                        {"award_id": "FA805122F0004", "obligated": 2000},
                    ],
                    "by_transaction": [
                        {"action_date": "2025-01-01", "amount_usd": 300},
                    ],
                },
                "insights": {"headline": "source prose, not a sheet"},
            }
        ),
        encoding="utf-8",
    )

    proc = subprocess.run(
        [
            sys.executable,
            str(RENDERERS / "render_xlsx.py"),
            "--input",
            str(source),
            "--output",
            str(output),
            "--title",
            "Competitive Intel Workbook",
        ],
        text=True,
        capture_output=True,
        check=False,
    )

    assert proc.returncode == 0, proc.stderr
    assert output.is_file()
    from openpyxl import load_workbook

    workbook = load_workbook(output)
    assert set(workbook.sheetnames) == {"obligations_by_award", "obligations_by_transaction"}


def test_render_docx_fallback_without_pandoc(tmp_path: Path) -> None:
    source = tmp_path / "brief.md"
    output = tmp_path / "brief.docx"
    source.write_text("# Brief\n\n- Finding one\n\nPlain paragraph.", encoding="utf-8")
    env = os.environ.copy()
    env["PATH"] = ""

    proc = subprocess.run(
        [
            sys.executable,
            str(RENDERERS / "render_docx.py"),
            "--input",
            str(source),
            "--output",
            str(output),
            "--metadata",
            "title=Brief",
        ],
        text=True,
        capture_output=True,
        check=False,
        env=env,
    )

    assert proc.returncode == 0, proc.stderr
    assert output.is_file()
    with zipfile.ZipFile(output) as docx:
        names = set(docx.namelist())
        assert "word/document.xml" in names
        document_xml = docx.read("word/document.xml").decode("utf-8")
    assert "Finding one" in document_xml